#!/usr/bin/env python3
"""Finanzas PC App — Genera un Excel maestro completo (por año) desde finanzas-data.json en Google Drive."""

import json
import os
import shutil
import sys
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter as gcl
    from openpyxl.chart import BarChart, Reference
except ImportError:
    print("Instalando openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter as gcl
    from openpyxl.chart import BarChart, Reference

MONTHS = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
          "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
SYNC_FILE = "finanzas-data.json"
SYNC_FOLDER = "Finanzas"
EXCEL_FOLDER = "Excel"
# Copia local del Excel maestro, para poder consultarlo sin depender de que
# la app de escritorio de Google Drive esté abierta y sincronizada
LOCAL_EXCEL_DIR = Path.home() / "Desktop" / "Fivvo Excel"

# Estilos
BOLD = Font(name="Calibri", size=10, bold=True)
NORM = Font(name="Calibri", size=10)
HDR_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HDR_FILL = PatternFill(start_color="14140F", end_color="14140F", fill_type="solid")
ACCENT_FILL = PatternFill(start_color="F5F5EB", end_color="F5F5EB", fill_type="solid")
POS_FONT = Font(name="Calibri", size=10, bold=True, color="22C55E")
NEG_FONT = Font(name="Calibri", size=10, bold=True, color="EF4444")
INV_FONT = Font(name="Calibri", size=10, bold=True, color="7C6EF6")
ITALIC_GRAY = Font(name="Calibri", size=9, italic=True, color="999999")
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(bottom=THIN)
EUR = '#,##0.00 "€"'
PCT = "0.0%"


def find_drive_folder():
    """Auto-detecta la carpeta local de Google Drive."""
    home = Path.home()
    candidates = [
        home / "Google Drive" / "My Drive",
        home / "Google Drive",
        home / "Mi unidad",
        Path("G:/My Drive"),
        Path("G:/Mi unidad"),
    ]
    for p in candidates:
        sync = p / SYNC_FOLDER / SYNC_FILE
        if sync.exists():
            return p / SYNC_FOLDER
    if sys.platform == "win32":
        for letter in "CDEFGHIJKLMNOPQRSTUVWXYZ":
            for sub in ["Google Drive/My Drive", "Google Drive/Mi unidad", "My Drive", "Mi unidad"]:
                p = Path(f"{letter}:/{sub}/{SYNC_FOLDER}/{SYNC_FILE}")
                if p.exists():
                    return p.parent
    return None


def read_sync_data(json_path):
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if data.get("app") != "finanzas-personales":
        raise ValueError("Archivo no es de Finanzas Personales")
    return data


def get_years_with_data(data):
    years = set()
    for tx in data.get("transactions", []):
        if tx.get("deleted"):
            continue
        if tx.get("year") is not None:
            years.add(tx["year"])
    for v in data.get("values", []):
        vid = v.get("id", "")
        if "-" in vid:
            try:
                years.add(int(vid.split("-")[0]))
            except (ValueError, IndexError):
                pass
    return sorted(years)


def W(ws, r, c, val=None, font=None, nf=None, fill=None, align="left", border=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = font or NORM
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if nf:
        cell.number_format = nf
    if fill:
        cell.fill = fill
    if border:
        cell.border = border
    return cell


def compute_month(year, month, data):
    """Calcula todos los totales de un mes concreto a partir de los datos sincronizados."""
    config = data.get("config", {})
    banks = config.get("banks", [])
    inv_types = config.get("inv_types", [])

    txs = [t for t in data.get("transactions", [])
           if t.get("year") == year and t.get("month") == month and not t.get("deleted")]
    ingresos = sorted([t for t in txs if t.get("type") == "ingreso"], key=lambda t: t.get("date", ""))
    gastos = sorted([t for t in txs if t.get("type") == "gasto"], key=lambda t: t.get("date", ""))
    inversiones = sorted([t for t in txs if t.get("type") == "inversion"], key=lambda t: t.get("date", ""))

    val_id = f"{year}-{month}"
    values = next((v for v in data.get("values", []) if v.get("id") == val_id), {})
    bancos = values.get("bancos", {})
    inv_ap = values.get("inv_aportado", {})
    inv_re = values.get("inv_valor_real", {})

    total_inc = sum(t.get("amount", 0) for t in ingresos)
    total_gas = sum(t.get("amount", 0) for t in gastos)
    total_inv = sum(t.get("amount", 0) for t in inversiones)
    balance = total_inc - total_gas
    tasa = (balance / total_inc) if total_inc > 0 else 0

    total_liq = sum((bancos.get(b, 0) or 0) for b in banks)
    total_ap = sum((inv_ap.get(t, 0) or 0) for t in inv_types)
    total_re = sum((inv_re.get(t, 0) or 0) for t in inv_types)
    patrimonio = total_liq + total_re

    return dict(ingresos=ingresos, gastos=gastos, inversiones=inversiones,
                total_inc=total_inc, total_gas=total_gas, total_inv=total_inv,
                balance=balance, tasa=tasa, bancos=bancos, inv_ap=inv_ap, inv_re=inv_re,
                total_liq=total_liq, total_ap=total_ap, total_re=total_re, patrimonio=patrimonio)


# ═══════ HOJA: Configuracion ═══════
def sh_configuracion(wb, banks, inv_types, platforms, inc_cats, exp_cats):
    ws = wb.active
    ws.title = "Configuracion"
    W(ws, 1, 1, "CONFIGURACIÓN", font=Font(name="Calibri", size=14, bold=True))
    for c, lbl in [(1, "BANCOS / LIQUIDEZ"), (2, "INVERSIONES"), (4, "PLATAFORMAS"),
                   (6, "CAT. INGRESOS"), (7, "CAT. GASTOS")]:
        W(ws, 3, c, lbl, font=HDR_FONT, fill=HDR_FILL)
    maxn = max(len(banks), len(inv_types), len(platforms), len(inc_cats), len(exp_cats), 1)
    for i in range(maxn):
        r = 4 + i
        if i < len(banks): W(ws, r, 1, banks[i])
        if i < len(inv_types): W(ws, r, 2, inv_types[i])
        if i < len(platforms): W(ws, r, 4, platforms[i])
        if i < len(inc_cats): W(ws, r, 6, inc_cats[i])
        if i < len(exp_cats): W(ws, r, 7, exp_cats[i])
    for c, w in [(1, 24), (2, 28), (3, 2), (4, 18), (5, 2), (6, 18), (7, 18)]:
        ws.column_dimensions[gcl(c)].width = w
    return ws


# ═══════ HOJA: PatrimonioTotal ═══════
def sh_patrimonio(wb, year, months_data, banks, inv_types):
    ws = wb.create_sheet("PatrimonioTotal")
    W(ws, 1, 1, "PATRIMONIO TOTAL", font=Font(name="Calibri", size=14, bold=True))
    W(ws, 3, 1, "CONCEPTO", font=HDR_FONT, fill=HDR_FILL)
    for i, mn in enumerate(MONTHS):
        W(ws, 3, 2 + i, mn, font=HDR_FONT, fill=HDR_FILL, align="center")

    row = 5
    W(ws, row, 1, "LIQUIDEZ", font=BOLD); row += 1
    for b in banks:
        W(ws, row, 1, b)
        for mi in range(12):
            W(ws, row, 2 + mi, (months_data[mi]["bancos"].get(b, 0) or 0), nf=EUR)
        row += 1
    W(ws, row, 1, "TOTAL LIQUIDEZ", font=BOLD)
    for mi in range(12):
        W(ws, row, 2 + mi, months_data[mi]["total_liq"], font=BOLD, nf=EUR)
    row += 2

    W(ws, row, 1, "INV. APORTADO", font=BOLD); row += 1
    for t in inv_types:
        W(ws, row, 1, t)
        for mi in range(12):
            W(ws, row, 2 + mi, (months_data[mi]["inv_ap"].get(t, 0) or 0), nf=EUR)
        row += 1
    W(ws, row, 1, "TOTAL APORTADO", font=BOLD)
    for mi in range(12):
        W(ws, row, 2 + mi, months_data[mi]["total_ap"], font=BOLD, nf=EUR)
    row += 2

    W(ws, row, 1, "INV. VALOR REAL", font=BOLD); row += 1
    for t in inv_types:
        W(ws, row, 1, t)
        for mi in range(12):
            W(ws, row, 2 + mi, (months_data[mi]["inv_re"].get(t, 0) or 0), nf=EUR)
        row += 1
    W(ws, row, 1, "TOTAL VALOR REAL", font=BOLD)
    for mi in range(12):
        W(ws, row, 2 + mi, months_data[mi]["total_re"], font=BOLD, nf=EUR)
    row += 2

    W(ws, row, 1, "PATRIMONIO NETO", font=Font(name="Calibri", size=11, bold=True))
    for mi in range(12):
        W(ws, row, 2 + mi, months_data[mi]["patrimonio"], font=Font(name="Calibri", size=11, bold=True), nf=EUR)

    ws.column_dimensions["A"].width = 26
    for i in range(12): ws.column_dimensions[gcl(2 + i)].width = 13
    ws.freeze_panes = "B4"
    return ws


# ═══════ HOJA: Inversiones ═══════
def sh_inversiones(wb, year, months_data, inv_types):
    ws = wb.create_sheet("Inversiones")
    W(ws, 1, 1, "INVERSIONES", font=Font(name="Calibri", size=14, bold=True))
    last = next((m for m in reversed(months_data) if m["total_ap"] > 0 or m["total_re"] > 0), months_data[-1])

    W(ws, 3, 1, "Capital aportado (actual)", font=BOLD)
    W(ws, 3, 2, last["total_ap"], font=BOLD, nf=EUR)
    W(ws, 4, 1, "Valor de mercado (actual)", font=BOLD)
    W(ws, 4, 2, last["total_re"], font=BOLD, nf=EUR)
    rent_eur = last["total_re"] - last["total_ap"]
    rent_pct = (rent_eur / last["total_ap"]) if last["total_ap"] > 0 else 0
    W(ws, 5, 1, "Rentabilidad", font=BOLD)
    W(ws, 5, 2, rent_eur, font=BOLD, nf=EUR)
    W(ws, 5, 3, rent_pct, font=BOLD, nf=PCT)

    for c, h in enumerate(["FONDO", "APORTADO", "VALOR ACTUAL", "RENT. EUR", "RENT. %"], 1):
        W(ws, 7, c, h, font=HDR_FONT, fill=HDR_FILL)
    for i, t in enumerate(inv_types):
        r = 8 + i
        ap = last["inv_ap"].get(t, 0) or 0
        re = last["inv_re"].get(t, 0) or 0
        W(ws, r, 1, t, border=BORDER)
        W(ws, r, 2, ap, nf=EUR, border=BORDER)
        W(ws, r, 3, re, nf=EUR, border=BORDER)
        W(ws, r, 4, re - ap, nf=EUR, border=BORDER)
        W(ws, r, 5, ((re - ap) / ap) if ap > 0 else 0, nf=PCT, border=BORDER)

    evo_r = 8 + len(inv_types) + 2
    W(ws, evo_r, 1, "EVOLUCIÓN MENSUAL", font=BOLD)
    for c, h in enumerate(["MES", "APORTADO", "VALOR REAL", "RENTABILIDAD"], 1):
        W(ws, evo_r + 1, c, h, font=HDR_FONT, fill=HDR_FILL)
    for i, mn in enumerate(MONTHS):
        r = evo_r + 2 + i
        W(ws, r, 1, mn)
        W(ws, r, 2, months_data[i]["total_ap"], nf=EUR)
        W(ws, r, 3, months_data[i]["total_re"], nf=EUR)
        W(ws, r, 4, months_data[i]["total_re"] - months_data[i]["total_ap"], nf=EUR)

    for c, w in [(1, 28), (2, 16), (3, 16), (4, 14), (5, 12)]: ws.column_dimensions[gcl(c)].width = w
    return ws


# ═══════ HOJA: Objetivos ═══════
def sh_objetivos(wb, year, months_data, goals):
    ws = wb.create_sheet("Objetivos")
    W(ws, 1, 1, "OBJETIVOS", font=Font(name="Calibri", size=14, bold=True))
    current_pat = next((m["patrimonio"] for m in reversed(months_data) if m["patrimonio"] > 0), 0)
    W(ws, 3, 1, "Patrimonio actual", font=BOLD)
    W(ws, 3, 2, current_pat, font=BOLD, nf=EUR)

    for c, h in enumerate(["OBJETIVO", "TIPO", "FECHA LÍMITE", "IMPORTE META", "% COMPLETADO"], 1):
        W(ws, 5, c, h, font=HDR_FONT, fill=HDR_FILL)
    if goals:
        for i, g in enumerate(goals):
            r = 6 + i
            target = g.get("target", 0) or 0
            pct = (current_pat / target) if target > 0 else 0
            W(ws, r, 1, g.get("name", ""), border=BORDER)
            W(ws, r, 2, g.get("type", ""), border=BORDER)
            W(ws, r, 3, g.get("deadline") or "", border=BORDER)
            W(ws, r, 4, target, nf=EUR, border=BORDER)
            W(ws, r, 5, pct, nf=PCT, border=BORDER)
    else:
        W(ws, 6, 1, "Sin objetivos definidos todavía", font=ITALIC_GRAY)

    cc, ch = 8, 5
    W(ws, ch, cc, "Mes", font=HDR_FONT, fill=HDR_FILL)
    W(ws, ch, cc + 1, "Patrimonio", font=HDR_FONT, fill=HDR_FILL)
    for i, mn in enumerate(MONTHS):
        r = ch + 1 + i
        W(ws, r, cc, mn)
        W(ws, r, cc + 1, months_data[i]["patrimonio"], nf=EUR)

    chart = BarChart()
    chart.type = "col"; chart.title = f"Evolución Patrimonio {year}"
    chart.y_axis.title = "EUR"; chart.width = 18; chart.height = 11
    chart.add_data(Reference(ws, min_col=cc + 1, max_col=cc + 1, min_row=ch, max_row=ch + 12), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=cc, max_col=cc, min_row=ch + 1, max_row=ch + 12))
    ws.add_chart(chart, "A15")

    for c, w in [(1, 26), (2, 12), (3, 14), (4, 16), (5, 14), (8, 10), (9, 14)]: ws.column_dimensions[gcl(c)].width = w
    return ws


# ═══════ HOJA: Presupuesto ═══════
def sh_presupuesto(wb, year, months_data, exp_cats, budget):
    ws = wb.create_sheet("Presupuesto")
    W(ws, 1, 1, f"PRESUPUESTO {year}", font=Font(name="Calibri", size=14, bold=True))

    total_limit = budget.get("total", 0) or 0
    real_anual_total = sum(m["total_gas"] for m in months_data)
    avg_monthly = real_anual_total / 12
    W(ws, 3, 1, "Presupuesto total mensual", font=BOLD)
    W(ws, 3, 2, total_limit, font=BOLD, nf=EUR)
    W(ws, 4, 1, "Gasto medio mensual (real)", font=BOLD)
    W(ws, 4, 2, avg_monthly, font=BOLD, nf=EUR)

    for c, h in enumerate(["CATEGORÍA", "LÍMITE MENSUAL", "LÍMITE ANUAL", "REAL ANUAL", "DIFERENCIA"], 1):
        W(ws, 6, c, h, font=HDR_FONT, fill=HDR_FILL)
    limits = budget.get("gastos", {})
    for i, cat in enumerate(exp_cats):
        r = 7 + i
        lim = limits.get(cat, 0) or 0
        real_anual = sum(sum(t["amount"] for t in m["gastos"] if t.get("category") == cat) for m in months_data)
        over = lim > 0 and real_anual > lim * 12
        W(ws, r, 1, cat, border=BORDER)
        W(ws, r, 2, lim, nf=EUR, border=BORDER)
        W(ws, r, 3, lim * 12, nf=EUR, border=BORDER)
        W(ws, r, 4, real_anual, nf=EUR, border=BORDER, font=NEG_FONT if over else NORM)
        W(ws, r, 5, lim * 12 - real_anual, nf=EUR, border=BORDER)

    for c, w in [(1, 24), (2, 16), (3, 16), (4, 16), (5, 16)]: ws.column_dimensions[gcl(c)].width = w
    return ws


# ═══════ HOJA MES (una por mes) ═══════
def write_tx_table(ws, row, title, txs, bank_field, bank_label, color):
    W(ws, row, 1, title, font=HDR_FONT, fill=HDR_FILL); row += 1
    for c, h in enumerate(["Fecha", bank_label, "Categoría", "Descripción", "Importe"], 1):
        W(ws, row, c, h, font=BOLD, border=BORDER)
    row += 1
    for t in txs:
        W(ws, row, 1, t.get("date", ""), border=BORDER)
        W(ws, row, 2, t.get(bank_field, ""), border=BORDER)
        W(ws, row, 3, t.get("category", ""), border=BORDER)
        W(ws, row, 4, t.get("description", ""), border=BORDER)
        W(ws, row, 5, t.get("amount", 0), nf=EUR, align="right", border=BORDER, font=color)
        row += 1
    if txs:
        W(ws, row, 4, "TOTAL", font=BOLD)
        W(ws, row, 5, sum(t.get("amount", 0) for t in txs), font=BOLD, nf=EUR, align="right")
    else:
        W(ws, row, 1, "Sin movimientos", font=ITALIC_GRAY)
    return row + 2


def sh_mes(wb, year, month, md, banks, inv_types):
    ws = wb.create_sheet(MONTHS[month])
    ws.sheet_properties.tabColor = "14140F"
    W(ws, 1, 1, f"{MONTHS[month].upper()} {year}", font=Font(name="Calibri", size=14, bold=True))

    row = 3
    for lbl, val, fmt, ft in [
        ("Ingresos", md["total_inc"], EUR, POS_FONT),
        ("Gastos", md["total_gas"], EUR, NEG_FONT),
        ("Inversiones", md["total_inv"], EUR, INV_FONT),
        ("Balance", md["balance"], EUR, POS_FONT if md["balance"] >= 0 else NEG_FONT),
        ("Tasa de ahorro", md["tasa"], PCT, BOLD),
        ("Patrimonio", md["patrimonio"], EUR, BOLD),
    ]:
        W(ws, row, 1, lbl, font=BOLD)
        W(ws, row, 2, val, font=ft, nf=fmt, align="right")
        row += 1
    row += 1

    W(ws, row, 1, "LIQUIDEZ", font=HDR_FONT, fill=HDR_FILL)
    W(ws, row, 2, "Saldo", font=HDR_FONT, fill=HDR_FILL, align="right"); row += 1
    for b in banks:
        W(ws, row, 1, b, border=BORDER)
        W(ws, row, 2, (md["bancos"].get(b, 0) or 0), nf=EUR, align="right", border=BORDER)
        row += 1
    W(ws, row, 1, "Total liquidez", font=BOLD)
    W(ws, row, 2, md["total_liq"], font=BOLD, nf=EUR, align="right"); row += 2

    W(ws, row, 1, "INVERSIONES", font=HDR_FONT, fill=HDR_FILL)
    W(ws, row, 2, "Aportado", font=HDR_FONT, fill=HDR_FILL, align="right")
    W(ws, row, 3, "Valor real", font=HDR_FONT, fill=HDR_FILL, align="right"); row += 1
    for t in inv_types:
        ap = md["inv_ap"].get(t, 0) or 0
        re = md["inv_re"].get(t, 0) or 0
        W(ws, row, 1, t, border=BORDER)
        W(ws, row, 2, ap, nf=EUR, align="right", border=BORDER)
        W(ws, row, 3, re, nf=EUR, align="right", border=BORDER)
        row += 1
    W(ws, row, 1, "Total inversiones", font=BOLD)
    W(ws, row, 2, md["total_ap"], font=BOLD, nf=EUR, align="right")
    W(ws, row, 3, md["total_re"], font=BOLD, nf=EUR, align="right")
    row += 3

    row = write_tx_table(ws, row, "INGRESOS", md["ingresos"], "bank", "Banco", POS_FONT)
    row = write_tx_table(ws, row, "GASTOS", md["gastos"], "bank", "Banco", NEG_FONT)
    row = write_tx_table(ws, row, "INVERSIONES", md["inversiones"], "platform", "Plataforma", INV_FONT)

    ws.freeze_panes = "A2"
    for c, w in [(1, 13), (2, 20), (3, 20), (4, 28), (5, 14)]: ws.column_dimensions[gcl(c)].width = w
    return ws


# ═══════ HOJA: ResumenAnual ═══════
def sh_resumen_anual(wb, year, months_data, inc_cats, exp_cats):
    ws = wb.create_sheet("ResumenAnual")
    W(ws, 1, 1, f"RESUMEN ANUAL {year}", font=Font(name="Calibri", size=14, bold=True))

    total_inc = sum(m["total_inc"] for m in months_data)
    total_gas = sum(m["total_gas"] for m in months_data)
    total_inv = sum(m["total_inv"] for m in months_data)
    balance = total_inc - total_gas
    tasa = (balance / total_inc) if total_inc > 0 else 0
    pat_dic = next((m["patrimonio"] for m in reversed(months_data) if m["patrimonio"] > 0), 0)

    for r, (lbl, val, fmt) in enumerate([
        ("Total ingresos", total_inc, EUR), ("Total gastos", total_gas, EUR),
        ("Total inversiones", total_inv, EUR), ("Balance del año", balance, EUR),
        ("Tasa de ahorro media", tasa, PCT), ("Patrimonio final", pat_dic, EUR),
    ], start=3):
        W(ws, r, 1, lbl, font=BOLD)
        W(ws, r, 2, val, font=BOLD, nf=fmt, align="right")

    hr = 10
    for c, h in enumerate(["MES", "INGRESOS", "GASTOS", "INVERSIONES", "BALANCE", "TASA AHORRO", "PATRIMONIO"], 1):
        W(ws, hr, c, h, font=HDR_FONT, fill=HDR_FILL)
    for i, mn in enumerate(MONTHS):
        r = hr + 1 + i; m = months_data[i]
        W(ws, r, 1, mn, border=BORDER)
        W(ws, r, 2, m["total_inc"], nf=EUR, border=BORDER)
        W(ws, r, 3, m["total_gas"], nf=EUR, border=BORDER)
        W(ws, r, 4, m["total_inv"], nf=EUR, border=BORDER)
        W(ws, r, 5, m["balance"], nf=EUR, border=BORDER)
        W(ws, r, 6, m["tasa"], nf=PCT, border=BORDER)
        W(ws, r, 7, m["patrimonio"], nf=EUR, border=BORDER)

    gx = hr + 14
    W(ws, gx, 1, "GASTOS POR CATEGORÍA (anual)", font=HDR_FONT, fill=HDR_FILL)
    W(ws, gx, 2, "Importe", font=HDR_FONT, fill=HDR_FILL, align="right")
    for i, cat in enumerate(exp_cats):
        r = gx + 1 + i
        total = sum(sum(t["amount"] for t in m["gastos"] if t.get("category") == cat) for m in months_data)
        W(ws, r, 1, cat, border=BORDER)
        W(ws, r, 2, total, nf=EUR, align="right", border=BORDER)

    for c, w in [(1, 20), (2, 16), (3, 16), (4, 16), (5, 16), (6, 14), (7, 16)]: ws.column_dimensions[gcl(c)].width = w
    return ws


# ═══════ HOJA: Dashboard ═══════
def sh_dashboard(wb, year, months_data, banks, inv_types, goals):
    ws = wb.create_sheet("Dashboard")
    active_idx = None
    for i, m in enumerate(months_data):
        if m["total_inc"] > 0 or m["total_gas"] > 0 or m["total_inv"] > 0 or m["patrimonio"] > 0:
            active_idx = i
    md = months_data[active_idx] if active_idx is not None else months_data[0]
    active_name = MONTHS[active_idx] if active_idx is not None else MONTHS[0]

    W(ws, 1, 1, f"DASHBOARD — FINANZAS {year}", font=Font(name="Calibri", size=16, bold=True))
    W(ws, 2, 1, f"Mes activo: {active_name}", font=Font(name="Calibri", size=10, color="999999"))

    W(ws, 4, 1, "PATRIMONIO", font=HDR_FONT, fill=HDR_FILL)
    W(ws, 5, 1, "Patrimonio neto total", font=BOLD)
    W(ws, 5, 2, md["patrimonio"], font=BOLD, nf=EUR, align="right")
    W(ws, 6, 1, "Liquidez")
    W(ws, 6, 2, md["total_liq"], nf=EUR, align="right")
    W(ws, 7, 1, "Inversiones — valor real")
    W(ws, 7, 2, md["total_re"], nf=EUR, align="right")
    rent_eur = md["total_re"] - md["total_ap"]
    rent_pct = (rent_eur / md["total_ap"]) if md["total_ap"] > 0 else 0
    W(ws, 8, 1, "Inversiones — rentabilidad")
    W(ws, 8, 2, rent_eur, nf=EUR, align="right")
    W(ws, 8, 3, rent_pct, nf=PCT)

    W(ws, 10, 1, "ESTE MES", font=HDR_FONT, fill=HDR_FILL)
    W(ws, 11, 1, "Ingresos")
    W(ws, 11, 2, md["total_inc"], nf=EUR, align="right", font=POS_FONT)
    W(ws, 12, 1, "Gastos")
    W(ws, 12, 2, md["total_gas"], nf=EUR, align="right", font=NEG_FONT)
    W(ws, 13, 1, "Balance", font=BOLD)
    W(ws, 13, 2, md["balance"], font=BOLD, nf=EUR, align="right")
    W(ws, 14, 1, "Tasa de ahorro", font=BOLD)
    W(ws, 14, 2, md["tasa"], font=BOLD, nf=PCT, align="right")

    r = 16
    W(ws, r, 1, "LIQUIDEZ POR CUENTA", font=HDR_FONT, fill=HDR_FILL); r += 1
    for b in banks:
        W(ws, r, 1, b)
        W(ws, r, 2, (md["bancos"].get(b, 0) or 0), nf=EUR, align="right")
        r += 1
    W(ws, r, 1, "Total", font=BOLD)
    W(ws, r, 2, md["total_liq"], font=BOLD, nf=EUR, align="right")
    r += 3

    W(ws, r, 1, "FONDO DE EMERGENCIA", font=HDR_FONT, fill=HDR_FILL); r += 1
    meses_cubiertos = (md["total_liq"] / md["total_gas"]) if md["total_gas"] > 0 else 0
    W(ws, r, 1, "Meses de gastos cubiertos", font=BOLD)
    W(ws, r, 2, meses_cubiertos, font=BOLD, nf="0.0", align="right")
    W(ws, r, 3, "(recomendado: mínimo 3)", font=Font(name="Calibri", size=9, color="999999"))
    r += 3

    if goals:
        W(ws, r, 1, "OBJETIVOS", font=HDR_FONT, fill=HDR_FILL); r += 1
        for g in goals[:3]:
            target = g.get("target", 0) or 0
            pct = (md["patrimonio"] / target) if target > 0 else 0
            W(ws, r, 1, g.get("name", ""))
            W(ws, r, 2, target, nf=EUR, align="right")
            W(ws, r, 3, pct, nf=PCT)
            r += 1

    cc, ch = 6, 4
    W(ws, ch, cc, "Mes", font=HDR_FONT, fill=HDR_FILL)
    W(ws, ch, cc + 1, "Patrimonio", font=HDR_FONT, fill=HDR_FILL)
    for i, mn in enumerate(MONTHS):
        rr = ch + 1 + i
        W(ws, rr, cc, mn)
        W(ws, rr, cc + 1, months_data[i]["patrimonio"], nf=EUR)

    chart = BarChart()
    chart.type = "col"; chart.title = f"Evolución del patrimonio {year}"
    chart.y_axis.title = "EUR"; chart.width = 20; chart.height = 14
    chart.add_data(Reference(ws, min_col=cc + 1, max_col=cc + 1, min_row=ch, max_row=ch + 12), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=cc, max_col=cc, min_row=ch + 1, max_row=ch + 12))
    ws.add_chart(chart, "A22")

    for c, w in [(1, 26), (2, 16), (3, 20), (6, 10), (7, 14)]: ws.column_dimensions[gcl(c)].width = w
    ws.freeze_panes = "A3"
    return ws


# ═══════ CONSTRUCCIÓN DEL LIBRO MAESTRO ═══════
def build_master_workbook(year, data, output_path):
    config = data.get("config", {})
    banks = config.get("banks", [])
    platforms = config.get("platforms", [])
    inv_types = config.get("inv_types", [])
    inc_cats = config.get("inc_categories", [])
    exp_cats = config.get("exp_categories", [])

    months_data = [compute_month(year, m, data) for m in range(12)]
    goals = [g for g in data.get("goals", []) if not g.get("deleted")]
    budget = next((b for b in data.get("budgets", []) if b.get("id") == "main"), {})

    wb = openpyxl.Workbook()
    sh_configuracion(wb, banks, inv_types, platforms, inc_cats, exp_cats)
    sh_patrimonio(wb, year, months_data, banks, inv_types)
    sh_inversiones(wb, year, months_data, inv_types)
    sh_objetivos(wb, year, months_data, goals)
    sh_presupuesto(wb, year, months_data, exp_cats, budget)
    for m in range(12):
        sh_mes(wb, year, m, months_data[m], banks, inv_types)
    sh_resumen_anual(wb, year, months_data, inc_cats, exp_cats)
    sh_dashboard(wb, year, months_data, banks, inv_types, goals)
    wb.move_sheet("Dashboard", offset=-wb.sheetnames.index("Dashboard"))

    wb.save(output_path)
    return True


def sync_local_copy(filepath, filename):
    """Copia el maestro a LOCAL_EXCEL_DIR, sustituyendo la copia anterior.
    Permite consultarlo sin depender de que Google Drive este sincronizando."""
    try:
        LOCAL_EXCEL_DIR.mkdir(parents=True, exist_ok=True)
        local_path = LOCAL_EXCEL_DIR / filename
        if filepath.exists() and (not local_path.exists() or local_path.stat().st_mtime < filepath.stat().st_mtime):
            shutil.copy2(filepath, local_path)
    except Exception as e:
        print(f"  Aviso: no se pudo copiar {filename} a la carpeta local: {e}")


def check_and_generate(drive_path):
    """Revisa cambios y regenera el/los Excel maestros (uno por año con datos).
    Cada maestro se copia ademas a LOCAL_EXCEL_DIR, reemplazando la version anterior."""
    json_path = drive_path / SYNC_FILE
    if not json_path.exists():
        print(f"No se encuentra {SYNC_FILE} en {drive_path}")
        return 0

    excel_dir = drive_path / EXCEL_FOLDER
    excel_dir.mkdir(exist_ok=True)

    data = read_sync_data(json_path)
    json_mtime = json_path.stat().st_mtime
    years = get_years_with_data(data) or [datetime.now().year]
    generated = 0

    for year in years:
        filename = f"Plantilla_Finanzas_{year}.xlsx"
        filepath = excel_dir / filename
        needs_update = not filepath.exists() or filepath.stat().st_mtime < json_mtime
        if needs_update:
            try:
                build_master_workbook(year, data, filepath)
                print(f"  Generado: {filename}")
                generated += 1
            except Exception as e:
                print(f"  Error generando {filename}: {e}")
                continue
        sync_local_copy(filepath, filename)

    return generated


def run_gui():
    try:
        import tkinter as tk
        from tkinter import filedialog, scrolledtext
    except ImportError:
        print("tkinter no disponible. Ejecutando en modo consola.")
        run_console()
        return

    drive_path = find_drive_folder()

    root = tk.Tk()
    root.title("Finanzas — Generador de Excel")
    root.geometry("550x400")
    root.configure(bg="#f5f5eb")

    frame = tk.Frame(root, bg="#f5f5eb", padx=20, pady=20)
    frame.pack(fill=tk.BOTH, expand=True)

    tk.Label(frame, text="Finanzas — Excel", font=("Calibri", 18, "bold"),
             bg="#f5f5eb", fg="#14140f").pack(anchor="w")

    tk.Label(frame, text="Genera un Excel maestro completo desde Google Drive",
             font=("Calibri", 11), bg="#f5f5eb", fg="#6e6e64").pack(anchor="w", pady=(0, 12))

    path_frame = tk.Frame(frame, bg="#f5f5eb")
    path_frame.pack(fill=tk.X, pady=(0, 8))

    tk.Label(path_frame, text="Carpeta Finanzas:", font=("Calibri", 10, "bold"),
             bg="#f5f5eb", fg="#14140f").pack(side=tk.LEFT)

    path_var = tk.StringVar(value=str(drive_path) if drive_path else "No encontrada")
    path_entry = tk.Entry(path_frame, textvariable=path_var, font=("Calibri", 10),
                          bg="white", fg="#14140f", width=35)
    path_entry.pack(side=tk.LEFT, padx=(8, 4))

    def browse():
        d = filedialog.askdirectory(title="Selecciona la carpeta Finanzas en Google Drive")
        if d:
            path_var.set(d)

    tk.Button(path_frame, text="...", command=browse, font=("Calibri", 10),
              bg="#14140f", fg="white", width=3).pack(side=tk.LEFT)

    log = scrolledtext.ScrolledText(frame, height=12, font=("Consolas", 9),
                                     bg="white", fg="#14140f", wrap=tk.WORD)
    log.pack(fill=tk.BOTH, expand=True, pady=(8, 8))

    def log_msg(msg):
        log.insert(tk.END, f"[{datetime.now().strftime('%H:%M:%S')}] {msg}\n")
        log.see(tk.END)
        root.update_idletasks()

    def do_generate():
        p = Path(path_var.get())
        if not p.exists():
            log_msg(f"ERROR: Carpeta no encontrada: {p}")
            return
        json_path = p / SYNC_FILE
        if not json_path.exists():
            log_msg(f"ERROR: No se encuentra {SYNC_FILE} en {p}")
            return

        log_msg("Leyendo datos...")
        try:
            data = read_sync_data(json_path)
        except Exception as e:
            log_msg(f"ERROR: {e}")
            return

        years = get_years_with_data(data) or [datetime.now().year]
        log_msg(f"Años con datos: {', '.join(str(y) for y in years)}")

        excel_dir = p / EXCEL_FOLDER
        excel_dir.mkdir(exist_ok=True)
        json_mtime = json_path.stat().st_mtime
        generated = 0

        for year in years:
            filename = f"Plantilla_Finanzas_{year}.xlsx"
            filepath = excel_dir / filename
            needs_update = not filepath.exists() or filepath.stat().st_mtime < json_mtime
            if needs_update:
                try:
                    build_master_workbook(year, data, filepath)
                    log_msg(f"Generado: {filename}")
                    generated += 1
                except Exception as e:
                    log_msg(f"Error: {filename} — {e}")
                    continue
            sync_local_copy(filepath, filename)

        log_msg(f"Copia local en: {LOCAL_EXCEL_DIR}")
        if generated == 0:
            log_msg("El Excel ya está actualizado")
        else:
            log_msg(f"Completado: {generated} archivo{'s' if generated != 1 else ''} generado{'s' if generated != 1 else ''}")

    watching = [False]
    watch_after_id = [None]

    def do_watch():
        if watching[0]:
            watching[0] = False
            if watch_after_id[0]:
                root.after_cancel(watch_after_id[0])
                watch_after_id[0] = None
            watch_btn.config(text="Vigilar cambios", bg="#14140f")
            log_msg("Vigilancia detenida")
            return

        watching[0] = True
        watch_btn.config(text="Detener vigilancia", bg="#ef4444")
        log_msg("Vigilando cambios cada 30 segundos...")

        last_mtime = [0]

        def check():
            if not watching[0]:
                return
            p = Path(path_var.get())
            json_path = p / SYNC_FILE
            if json_path.exists():
                mtime = json_path.stat().st_mtime
                if mtime > last_mtime[0]:
                    last_mtime[0] = mtime
                    if mtime > 0:
                        log_msg("Cambio detectado, regenerando...")
                        do_generate()
            watch_after_id[0] = root.after(30000, check)

        check()

    btn_frame = tk.Frame(frame, bg="#f5f5eb")
    btn_frame.pack(fill=tk.X)

    tk.Button(btn_frame, text="Generar Excel", command=do_generate,
              font=("Calibri", 11, "bold"), bg="#beff50", fg="#14140f",
              padx=16, pady=6, cursor="hand2").pack(side=tk.LEFT, padx=(0, 8))

    watch_btn = tk.Button(btn_frame, text="Vigilar cambios", command=do_watch,
                          font=("Calibri", 11, "bold"), bg="#14140f", fg="white",
                          padx=16, pady=6, cursor="hand2")
    watch_btn.pack(side=tk.LEFT)

    if drive_path:
        log_msg(f"Carpeta detectada: {drive_path}")
    else:
        log_msg("Carpeta Google Drive no detectada. Selecciónala manualmente.")

    root.mainloop()


def run_console():
    drive_path = find_drive_folder()
    if not drive_path:
        print("No se encontró la carpeta Finanzas en Google Drive.")
        path = input("Introduce la ruta manualmente: ").strip()
        if not path:
            return
        drive_path = Path(path)

    print(f"Carpeta: {drive_path}")
    n = check_and_generate(drive_path)
    if n == 0:
        print("El Excel ya está actualizado.")
    else:
        print(f"Generados {n} archivo(s) Excel.")


if __name__ == "__main__":
    if "--console" in sys.argv:
        run_console()
    else:
        run_gui()
