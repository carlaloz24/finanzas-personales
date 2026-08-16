#!/usr/bin/env python3
"""
Importador de datos desde la PWA Finanzas al Excel.

Lee un archivo JSON exportado desde la app móvil e introduce cada registro
en la hoja y tabla correcta del Excel (Plantilla_Finanzas_2025_v5.xlsx).

Uso:
  - Doble clic en este archivo (si Python está asociado a .py)
  - O desde terminal: python importar_datos.py

Detecta duplicados por fecha+tipo+categoría+importe+descripción.
"""

import json
import sys
import os
import hashlib
from tkinter import Tk, filedialog, messagebox, Label, Button, Frame, StringVar, ttk
from tkinter import font as tkfont
import openpyxl

# ══════════════════════════════════════════════════════════════
# LAYOUT DEL EXCEL (debe coincidir con build_plantilla.py)
# ══════════════════════════════════════════════════════════════
MONTHS = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
          "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]

BANKS     = ["Banco Santander", "Revolut", "OpenBank", "MyInvestor", "Efectivo"]
INVS      = ["Fidelity MSCI World P-ACC-EUR", "Vanguard Small-Cap EUR Acc", "Vanguard Emerging Mkts EUR Acc", "Nueva Expresion Textil (NEXTIL)"]

NB = len(BANKS)
NI = len(INVS)

ING_HDR = 4;   ING_S = 5;   ING_E = 54
GAS_HDR = 57;  GAS_S = 58;  GAS_E = 107
INV_HDR = 110; INV_S = 111; INV_E = 140

VB_S = 3
VI_S = 14
VR_S = 23


def tx_hash(date, tx_type, category, amount, description):
    """Hash determinista para detectar duplicados."""
    raw = f"{date}|{tx_type}|{category}|{amount:.2f}|{description or ''}"
    return hashlib.sha256(raw.encode()).hexdigest()[:16]


def find_next_empty_row(ws, start_row, end_row, col=1):
    """Encuentra la primera fila vacía en el rango (col A = Fecha)."""
    for r in range(start_row, end_row + 1):
        if ws.cell(r, col).value is None:
            return r
    return None


def get_existing_hashes(ws, start_row, end_row, tx_type):
    """Lee las filas existentes y genera hashes para detectar duplicados."""
    hashes = set()
    for r in range(start_row, end_row + 1):
        fecha = ws.cell(r, 1).value
        if fecha is None:
            break
        cat = ws.cell(r, 3).value or ""
        imp = ws.cell(r, 5).value
        desc = ws.cell(r, 4).value or ""
        if imp is not None:
            date_str = str(fecha)[:10] if hasattr(fecha, 'strftime') else str(fecha)
            h = tx_hash(date_str, tx_type, cat, float(imp), desc)
            hashes.add(h)
    return hashes


def import_transactions(ws, txs, tx_type, start_row, end_row, existing_hashes):
    """Importa una lista de transacciones a la hoja del mes."""
    imported = 0
    skipped = 0
    errors = []

    for tx in txs:
        date = tx.get('date', '')
        amount = tx.get('amount', 0)
        desc = tx.get('description', '')

        if tx_type == 'inversion':
            cat = tx.get('tipo_inv', tx.get('category', ''))
            bank = tx.get('platform', '')
        else:
            cat = tx.get('category', '')
            bank = tx.get('bank', '')

        h = tx_hash(date, tx_type, cat, amount, desc)
        if h in existing_hashes:
            skipped += 1
            continue

        row = find_next_empty_row(ws, start_row, end_row)
        if row is None:
            errors.append(f"Sin espacio para {tx_type}: {cat} {amount}")
            continue

        ws.cell(row, 1).value = date
        ws.cell(row, 1).number_format = 'DD/MM/YYYY'
        ws.cell(row, 2).value = bank
        ws.cell(row, 3).value = cat
        ws.cell(row, 4).value = desc
        ws.cell(row, 5).value = round(amount, 2)
        ws.cell(row, 5).number_format = '#,##0.00 "EUR"'

        existing_hashes.add(h)
        imported += 1

    return imported, skipped, errors


def import_values(ws, valores):
    """Escribe los valores actuales (saldos, inversiones) en las celdas fijas."""
    if not valores:
        return

    bancos = valores.get('bancos', {})
    for name, val in bancos.items():
        if name in BANKS:
            idx = BANKS.index(name)
            ws.cell(VB_S + idx, 11).value = round(val, 2)
            ws.cell(VB_S + idx, 11).number_format = '#,##0.00 "EUR"'

    inv_ap = valores.get('inv_aportado', {})
    for name, val in inv_ap.items():
        if name in INVS:
            idx = INVS.index(name)
            ws.cell(VI_S + idx, 11).value = round(val, 2)
            ws.cell(VI_S + idx, 11).number_format = '#,##0.00 "EUR"'

    inv_real = valores.get('inv_valor_real', {})
    for name, val in inv_real.items():
        if name in INVS:
            idx = INVS.index(name)
            ws.cell(VR_S + idx, 11).value = round(val, 2)
            ws.cell(VR_S + idx, 11).number_format = '#,##0.00 "EUR"'


def process_month(wb, month_data):
    """Procesa los datos de un mes: transacciones + valores."""
    month_num = month_data.get('month')
    month_name = month_data.get('month_name', '')

    if month_name not in MONTHS:
        if isinstance(month_num, int) and 0 <= month_num <= 11:
            month_name = MONTHS[month_num]
        else:
            return None, f"Mes no reconocido: {month_name} / {month_num}"

    if month_name not in wb.sheetnames:
        return None, f"Hoja '{month_name}' no encontrada en el Excel"

    ws = wb[month_name]

    ing_hashes = get_existing_hashes(ws, ING_S, ING_E, 'ingreso')
    gas_hashes = get_existing_hashes(ws, GAS_S, GAS_E, 'gasto')
    inv_hashes = get_existing_hashes(ws, INV_S, INV_E, 'inversion')

    txs = month_data.get('transactions', month_data)
    ingresos = txs.get('ingresos', [])
    gastos = txs.get('gastos', [])
    inversiones = txs.get('inversiones', [])

    results = {'imported': 0, 'skipped': 0, 'errors': []}

    imp, skip, errs = import_transactions(ws, ingresos, 'ingreso', ING_S, ING_E, ing_hashes)
    results['imported'] += imp; results['skipped'] += skip; results['errors'] += errs

    imp, skip, errs = import_transactions(ws, gastos, 'gasto', GAS_S, GAS_E, gas_hashes)
    results['imported'] += imp; results['skipped'] += skip; results['errors'] += errs

    imp, skip, errs = import_transactions(ws, inversiones, 'inversion', INV_S, INV_E, inv_hashes)
    results['imported'] += imp; results['skipped'] += skip; results['errors'] += errs

    valores = month_data.get('valores_actuales')
    import_values(ws, valores)

    return month_name, results


def run_import(json_path, excel_path):
    """Ejecuta la importación completa."""
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    if data.get('app') != 'finanzas-personales':
        return False, "El archivo no es una exportación válida de Finanzas Personales."

    wb = openpyxl.load_workbook(excel_path)

    all_results = []

    if data.get('export_type') == 'full':
        for m in data.get('months', []):
            name, res = process_month(wb, m)
            if name:
                all_results.append((name, res))
            elif res:
                all_results.append(("Error", {'imported':0,'skipped':0,'errors':[res]}))
    else:
        name, res = process_month(wb, data)
        if name:
            all_results.append((name, res))
        elif res:
            all_results.append(("Error", {'imported':0,'skipped':0,'errors':[res]}))

    base, ext = os.path.splitext(excel_path)
    output_path = excel_path
    wb.save(output_path)

    total_imp = sum(r['imported'] for _, r in all_results)
    total_skip = sum(r['skipped'] for _, r in all_results)
    all_errors = []
    for _, r in all_results:
        all_errors += r['errors']

    report = []
    report.append(f"Importación completada")
    report.append(f"")
    for name, r in all_results:
        report.append(f"  {name}: {r['imported']} nuevos, {r['skipped']} duplicados omitidos")
    report.append(f"")
    report.append(f"Total: {total_imp} registros importados, {total_skip} duplicados omitidos")
    if all_errors:
        report.append(f"")
        report.append(f"Errores:")
        for e in all_errors:
            report.append(f"  - {e}")
    report.append(f"")
    report.append(f"Guardado en: {output_path}")

    return True, "\n".join(report)


# ══════════════════════════════════════════════════════════════
# GUI
# ══════════════════════════════════════════════════════════════
class ImporterApp:
    def __init__(self):
        self.root = Tk()
        self.root.title("Finanzas — Importador")
        self.root.geometry("560x420")
        self.root.configure(bg='#0f0f1a')
        self.root.resizable(False, False)

        try:
            self.root.iconbitmap(default='')
        except Exception:
            pass

        style = ttk.Style()
        style.theme_use('clam')

        self.json_path = StringVar(value="")
        self.excel_path = StringVar(value="")

        default_excel = self._find_default_excel()
        if default_excel:
            self.excel_path.set(default_excel)

        self._build_ui()

    def _find_default_excel(self):
        script_dir = os.path.dirname(os.path.abspath(__file__))
        for name in ["Plantilla_Finanzas_2025_v5.xlsx", "Plantilla_Finanzas_2025_v4.xlsx"]:
            path = os.path.join(script_dir, name)
            if os.path.exists(path):
                return path
        return ""

    def _build_ui(self):
        bg = '#0f0f1a'
        card = '#1a1a2e'
        accent = '#7c6ef0'
        text = '#f0f0ff'
        text2 = '#8888aa'

        title = Label(self.root, text="Importador de Finanzas", font=("Segoe UI", 18, "bold"),
                      bg=bg, fg=text)
        title.pack(pady=(24, 4))

        subtitle = Label(self.root, text="Importa datos del móvil al Excel",
                        font=("Segoe UI", 11), bg=bg, fg=text2)
        subtitle.pack(pady=(0, 20))

        json_frame = Frame(self.root, bg=card, highlightbackground='#2a2a45',
                          highlightthickness=1)
        json_frame.pack(padx=24, pady=6, fill='x')

        Label(json_frame, text="ARCHIVO JSON (exportado del móvil)",
              font=("Segoe UI", 9, "bold"), bg=card, fg=text2).pack(anchor='w', padx=12, pady=(10,0))
        json_inner = Frame(json_frame, bg=card)
        json_inner.pack(fill='x', padx=12, pady=(4,10))
        self.json_label = Label(json_inner, textvariable=self.json_path,
                               font=("Segoe UI", 10), bg=card, fg=text, anchor='w', width=38)
        self.json_label.pack(side='left', fill='x', expand=True)
        Button(json_inner, text="Buscar", font=("Segoe UI", 10, "bold"),
               bg=accent, fg='white', relief='flat', padx=12, pady=4,
               cursor='hand2', command=self._pick_json).pack(side='right')

        excel_frame = Frame(self.root, bg=card, highlightbackground='#2a2a45',
                           highlightthickness=1)
        excel_frame.pack(padx=24, pady=6, fill='x')

        Label(excel_frame, text="ARCHIVO EXCEL (plantilla destino)",
              font=("Segoe UI", 9, "bold"), bg=card, fg=text2).pack(anchor='w', padx=12, pady=(10,0))
        excel_inner = Frame(excel_frame, bg=card)
        excel_inner.pack(fill='x', padx=12, pady=(4,10))
        self.excel_label = Label(excel_inner, textvariable=self.excel_path,
                                font=("Segoe UI", 10), bg=card, fg=text, anchor='w', width=38)
        self.excel_label.pack(side='left', fill='x', expand=True)
        Button(excel_inner, text="Buscar", font=("Segoe UI", 10, "bold"),
               bg=accent, fg='white', relief='flat', padx=12, pady=4,
               cursor='hand2', command=self._pick_excel).pack(side='right')

        self.import_btn = Button(self.root, text="IMPORTAR",
                                font=("Segoe UI", 14, "bold"),
                                bg=accent, fg='white', relief='flat',
                                padx=40, pady=10, cursor='hand2',
                                command=self._do_import)
        self.import_btn.pack(pady=24)

        self.status = Label(self.root, text="Selecciona los archivos y pulsa Importar",
                           font=("Segoe UI", 10), bg=bg, fg=text2, wraplength=500,
                           justify='left')
        self.status.pack(padx=24, fill='x')

    def _pick_json(self):
        path = filedialog.askopenfilename(
            title="Selecciona el archivo JSON exportado",
            filetypes=[("JSON", "*.json"), ("Todos", "*.*")]
        )
        if path:
            self.json_path.set(path)

    def _pick_excel(self):
        path = filedialog.askopenfilename(
            title="Selecciona el Excel de destino",
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")]
        )
        if path:
            self.excel_path.set(path)

    def _do_import(self):
        jp = self.json_path.get()
        ep = self.excel_path.get()

        if not jp or not os.path.exists(jp):
            messagebox.showerror("Error", "Selecciona un archivo JSON válido.")
            return
        if not ep or not os.path.exists(ep):
            messagebox.showerror("Error", "Selecciona un archivo Excel válido.")
            return

        self.import_btn.config(state='disabled', text='Importando...')
        self.root.update()

        try:
            ok, report = run_import(jp, ep)
            if ok:
                self.status.config(text=report, fg='#00d4aa')
                messagebox.showinfo("Importación completada", report)
            else:
                self.status.config(text=report, fg='#ff6b6b')
                messagebox.showerror("Error", report)
        except Exception as e:
            msg = f"Error inesperado: {str(e)}"
            self.status.config(text=msg, fg='#ff6b6b')
            messagebox.showerror("Error", msg)
        finally:
            self.import_btn.config(state='normal', text='IMPORTAR')

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    app = ImporterApp()
    app.run()
