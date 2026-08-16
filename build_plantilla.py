#!/usr/bin/env python3
"""Plantilla Finanzas Personales 2025 v4 â€” Completa y funcional"""

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter as gcl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# DATOS REALES (Carla, junio 2025)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
BANKS     = ["Banco Santander", "Revolut", "OpenBank", "MyInvestor", "Efectivo"]
PLATFORMS = ["MyInvestor"]
INVS      = ["Fidelity MSCI World P-ACC-EUR", "Vanguard Small-Cap EUR Acc", "Vanguard Emerging Mkts EUR Acc", "Nueva Expresion Textil (NEXTIL)"]
INV_AMT   = [0, 0, 0, 43.71]
DEBTS     = []
INC_CAT   = ["Nomina", "Trabajo extra", "Ventas online", "Otros"]
EXP_CAT   = ["Comida", "Ocio", "Ropa", "Viajes", "Transporte",
             "Belleza", "Formacion", "Suscripciones", "Eventos", "Regalos", "Otros"]
MONTHS    = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
             "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
YEAR      = 2025

NB  = len(BANKS)     # 3
NP  = len(PLATFORMS) # 1
NI  = len(INVS)      # 4
NIC = len(INC_CAT)   # 3
NEC = len(EXP_CAT)   # 11
ND  = len(DEBTS)     # 0

# Saldos reales a 29/06/2025 (pre-rellena hoja Junio)
JUNIO_BANCOS = [1009.50, 0, 12517.88, 30.83, 0]
JUNIO_INV_AP = [0, 0, 0, 43.71]
JUNIO_INV_RE = [0, 0, 0, 43.71]

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# LAYOUT HOJAS MES
#
# Cols A(1)-E(5):   Transacciones [Fecha|Banco/Plat|Cat/Tipo|Desc|Importe]
# Col  F(6):        Separador
# Cols G(7)-H(8):   Resumen rapido + Desglose por categorias
# Col  I(9):        Separador
# Cols J(10)-K(11): Valores Actuales (filas FIJAS) -> PatrimonioTotal
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
ING_HDR = 4;   ING_S = 5;   ING_E = 54;  ING_TOT = 55
GAS_HDR = 57;  GAS_S = 58;  GAS_E = 107; GAS_TOT = 108
INV_HDR = 110; INV_S = 111; INV_E = 140; INV_TOT = 141

# Valores Actuales col K(11) â€” FILAS FIJAS en los 12 meses
VB_S = 3;  VB_E = VB_S + NB - 1   # bancos  K3:K5
VI_S = 14; VI_E = VI_S + NI - 1   # aportado K14:K17
VR_S = 23; VR_E = VR_S + NI - 1   # real     K23:K26
VT_LIQ = VB_E + 1                 # K6  total liquidez
VT_IA  = VI_E + 1                 # K18 total aportado
VT_IR  = VR_E + 1                 # K27 total real

# Posiciones fijas col H (para INDIRECT desde Dashboard / ResumenMensual)
H_ING = 5   # total ingresos
H_GAS = 6   # total gastos
H_INV = 7   # total inversiones
H_LIQ = 8   # total liquidez
H_BAL = 9   # balance
H_SAV = 10  # tasa ahorro

# Desglose G(label) / H(importe)
DSG_INC_S = 13
DSG_GAS_S = DSG_INC_S + NIC + 2   # = 18
DSG_INV_S = DSG_GAS_S + NEC + 2   # = 31

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# LAYOUT PatrimonioTotal
# Col A=etiquetas, B=Enero...M=Diciembre
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
P_LIQ_HDR = 5;  P_LIQ_S = 6;   P_LIQ_T = P_LIQ_S + NB    # 5,6,9
P_IA_HDR  = P_LIQ_T + 2        # 11
P_IA_S    = P_IA_HDR + 1       # 12
P_IA_T    = P_IA_S + NI        # 16
P_IR_HDR  = P_IA_T + 2         # 18
P_IR_S    = P_IR_HDR + 1       # 19
P_IR_T    = P_IR_S + NI        # 23

if ND > 0:
    P_DAH     = P_IR_T + 2
    P_DA_S    = P_DAH + 1
    P_DA_T    = P_DA_S + ND
    P_PAT_ROW = P_DA_T + 2
else:
    P_DAH = P_DA_S = P_DA_T = None
    P_PAT_ROW = P_IR_T + 2     # 25

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# HELPERS
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
BOLD = Font(name="Calibri", size=10, bold=True)
NORM = Font(name="Calibri", size=10)
EUR  = '#,##0.00 "EUR"'
PCT  = "0.00%"

def W(ws, r, c, val=None, f=None, bold=False, nf=None, align="left"):
    cell = ws.cell(row=r, column=c)
    if f is not None:   cell.value = f
    elif val is not None: cell.value = val
    cell.font = BOLD if bold else NORM
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if nf: cell.number_format = nf
    return cell

def CW(ws, c, w): ws.column_dimensions[gcl(c)].width = w

def add_dv(ws, col_n, r0, r1, formula):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True,
                        showErrorMessage=False, showInputMessage=False)
    ws.add_data_validation(dv)
    for r in range(r0, r1 + 1): dv.add(ws.cell(r, col_n))

CFG_BANKS = f"Configuracion!$A$4:$A${3+NB}"
CFG_PLATS = f"Configuracion!$L$4:$L${3+NP}"
CFG_INVS  = f"Configuracion!$B$4:$B${3+NI}"
CFG_INCAT = f"Configuracion!$I$4:$I${3+NIC}"
CFG_EXCAT = f"Configuracion!$J$4:$J${3+NEC}"


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# HOJA: Configuracion
# A=bancos, B=tipos inv, C=capital inv,
# E=deuda nombre, F=deuda, G=activo precio,
# I=cat ingresos, J=cat gastos, L=plataformas inv
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
def sh_configuracion(wb):
    ws = wb.active
    ws.title = "Configuracion"
    W(ws, 1, 1, f"CONFIGURACION {YEAR}", bold=True)
    for c, lbl in [(1,"BANCOS / LIQUIDEZ"),(2,"INVERSIONES"),(3,"CAPITAL INICIAL"),
                   (5,"DEUDA/ACTIVO"),(6,"DEUDA"),(7,"PRECIO ACTIVO"),
                   (9,"CAT. INGRESOS"),(10,"CAT. GASTOS"),(12,"PLATAFORMAS INV.")]:
        W(ws, 3, c, lbl, bold=True)
    for i in range(max(NB, NP, NI, NIC, NEC, max(ND, 1))):
        r = 4 + i
        if i < NB:  W(ws, r, 1, BANKS[i])
        if i < NI:  W(ws, r, 2, INVS[i]); W(ws, r, 3, INV_AMT[i], nf=EUR)
        if i < ND:
            nm, debt, asset = DEBTS[i]
            W(ws, r, 5, nm)
            if debt:  W(ws, r, 6, debt, nf=EUR)
            if asset: W(ws, r, 7, asset, nf=EUR)
        if i < NIC: W(ws, r, 9, INC_CAT[i])
        if i < NEC: W(ws, r, 10, EXP_CAT[i])
        if i < NP:  W(ws, r, 12, PLATFORMS[i])
    for c, w in [(1,20),(2,22),(3,15),(4,2),(5,18),(6,14),(7,14),(8,2),(9,20),(10,20),(11,2),(12,18)]:
        CW(ws, c, w)
    return ws


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# HOJA MES (funcion reutilizable para los 12 meses)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
def sh_mes(wb, month_name):
    ws = wb.create_sheet(month_name)
    W(ws, 1, 1, f"{month_name.upper()} {YEAR}", bold=True)

    # â”€â”€ RESUMEN RAPIDO (col G-H, filas 3-10) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    W(ws, 3,     7, "RESUMEN RAPIDO",    bold=True)
    W(ws, H_ING, 7, "Ingresos",          bold=True)
    W(ws, H_ING, 8, f=f"=IFERROR(SUM(E{ING_S}:E{ING_E}),0)",      nf=EUR, bold=True)
    W(ws, H_GAS, 7, "Gastos",            bold=True)
    W(ws, H_GAS, 8, f=f"=IFERROR(SUM(E{GAS_S}:E{GAS_E}),0)",      nf=EUR, bold=True)
    W(ws, H_INV, 7, "Inversiones",       bold=True)
    W(ws, H_INV, 8, f=f"=IFERROR(SUM(E{INV_S}:E{INV_E}),0)",      nf=EUR, bold=True)
    W(ws, H_LIQ, 7, "Liquidez total",    bold=True)
    W(ws, H_LIQ, 8, f=f"=IFERROR(SUM(K{VB_S}:K{VB_E}),0)",        nf=EUR, bold=True)
    W(ws, H_BAL, 7, "Balance",           bold=True)
    W(ws, H_BAL, 8, f=f"=IFERROR(H{H_ING}-H{H_GAS},0)",           nf=EUR, bold=True)
    W(ws, H_SAV, 7, "Tasa ahorro",       bold=True)
    W(ws, H_SAV, 8, f=f"=IFERROR((H{H_ING}-H{H_GAS})/H{H_ING},0)",nf=PCT, bold=True)

    # â”€â”€ DESGLOSE (col G-H, desde fila 13) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    W(ws, DSG_INC_S - 1, 7, "INGRESOS POR CATEGORIA", bold=True)
    for i in range(NIC):
        r = DSG_INC_S + i
        W(ws, r, 7, f=f"=Configuracion!I{4+i}")
        W(ws, r, 8, f=f"=IFERROR(SUMIF($C${ING_S}:$C${ING_E},G{r},$E${ING_S}:$E${ING_E}),0)", nf=EUR)

    W(ws, DSG_GAS_S - 1, 7, "GASTOS POR CATEGORIA", bold=True)
    for i in range(NEC):
        r = DSG_GAS_S + i
        W(ws, r, 7, f=f"=Configuracion!J{4+i}")
        W(ws, r, 8, f=f"=IFERROR(SUMIF($C${GAS_S}:$C${GAS_E},G{r},$E${GAS_S}:$E${GAS_E}),0)", nf=EUR)

    W(ws, DSG_INV_S - 1, 7, "INVERSIONES POR TIPO", bold=True)
    for i in range(NI):
        r = DSG_INV_S + i
        W(ws, r, 7, f=f"=Configuracion!B{4+i}")
        W(ws, r, 8, f=f"=IFERROR(SUMIF($C${INV_S}:$C${INV_E},G{r},$E${INV_S}:$E${INV_E}),0)", nf=EUR)

    # â”€â”€ VALORES ACTUALES (col J-K, FILAS FIJAS) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    W(ws, 1,  10, "VALORES ACTUALES",bold=True)
    W(ws, 2,  10, "LIQUIDEZ",        bold=True)
    for i in range(NB):
        r = VB_S + i
        W(ws, r, 10, f=f"=Configuracion!A{4+i}")
        W(ws, r, 11, 0, nf=EUR)
    W(ws, VT_LIQ, 10, "TOTAL LIQUIDEZ",  bold=True)
    W(ws, VT_LIQ, 11, f=f"=SUM(K{VB_S}:K{VB_E})", nf=EUR, bold=True)

    W(ws, 12, 10, "INV. APORTADO",   bold=True)
    for i in range(NI):
        r = VI_S + i
        W(ws, r, 10, f=f"=Configuracion!B{4+i}")
        W(ws, r, 11, 0, nf=EUR)
    W(ws, VT_IA, 10, "TOTAL APORTADO", bold=True)
    W(ws, VT_IA, 11, f=f"=SUM(K{VI_S}:K{VI_E})", nf=EUR, bold=True)

    W(ws, 21, 10, "INV. VALOR REAL", bold=True)
    for i in range(NI):
        r = VR_S + i
        W(ws, r, 10, f=f"=Configuracion!B{4+i}")
        W(ws, r, 11, 0, nf=EUR)
    W(ws, VT_IR, 10, "TOTAL REAL",     bold=True)
    W(ws, VT_IR, 11, f=f"=SUM(K{VR_S}:K{VR_E})", nf=EUR, bold=True)

    # â”€â”€ INGRESOS â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    W(ws, ING_HDR, 1, "-- INGRESOS --", bold=True)
    for c, lbl in enumerate(["Fecha","Banco","Categoria","Descripcion","Importe"], 1):
        W(ws, ING_HDR, c, lbl, bold=True)
    add_dv(ws, 2, ING_S, ING_E, CFG_BANKS)
    add_dv(ws, 3, ING_S, ING_E, CFG_INCAT)
    for r in range(ING_S, ING_E + 1): ws.cell(r, 5).number_format = EUR
    W(ws, ING_TOT, 4, "TOTAL INGRESOS", bold=True)
    W(ws, ING_TOT, 5, f=f"=SUM(E{ING_S}:E{ING_E})", nf=EUR, bold=True)

    # â”€â”€ GASTOS â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    W(ws, GAS_HDR, 1, "-- GASTOS --", bold=True)
    for c, lbl in enumerate(["Fecha","Banco","Categoria","Descripcion","Importe"], 1):
        W(ws, GAS_HDR, c, lbl, bold=True)
    add_dv(ws, 2, GAS_S, GAS_E, CFG_BANKS)
    add_dv(ws, 3, GAS_S, GAS_E, CFG_EXCAT)
    for r in range(GAS_S, GAS_E + 1): ws.cell(r, 5).number_format = EUR
    W(ws, GAS_TOT, 4, "TOTAL GASTOS", bold=True)
    W(ws, GAS_TOT, 5, f=f"=SUM(E{GAS_S}:E{GAS_E})", nf=EUR, bold=True)

    # â”€â”€ INVERSIONES â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    W(ws, INV_HDR, 1, "-- INVERSIONES --", bold=True)
    for c, lbl in enumerate(["Fecha","Plataforma","Tipo","Descripcion","Importe"], 1):
        W(ws, INV_HDR, c, lbl, bold=True)
    add_dv(ws, 2, INV_S, INV_E, CFG_PLATS)   # Plataforma = MyInvestor
    add_dv(ws, 3, INV_S, INV_E, CFG_INVS)    # Tipo = fondo/accion
    for r in range(INV_S, INV_E + 1): ws.cell(r, 5).number_format = EUR
    W(ws, INV_TOT, 4, "TOTAL INVERSIONES", bold=True)
    W(ws, INV_TOT, 5, f=f"=SUM(E{INV_S}:E{INV_E})", nf=EUR, bold=True)

    ws.freeze_panes = "A2"
    for c, w in [(1,11),(2,14),(3,18),(4,24),(5,13),(6,2),(7,26),(8,14),(9,2),(10,26),(11,14)]:
        CW(ws, c, w)
    return ws


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# HOJA: PatrimonioTotal
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
def sh_patrimonio(wb):
    ws = wb.create_sheet("PatrimonioTotal")
    W(ws, 1, 1, "PATRIMONIO TOTAL", bold=True)
    W(ws, 2, 1, "Fondo de Emergencia (editable):", bold=True)
    W(ws, 2, 2, 10000, nf=EUR, bold=True)
    W(ws, 3, 1, "PATRIMONIO ACTUAL:", bold=True)
    W(ws, 3, 2,
      f=f"=IFERROR(LOOKUP(2,1/(B{P_PAT_ROW}:M{P_PAT_ROW}>0),B{P_PAT_ROW}:M{P_PAT_ROW}),0)",
      nf=EUR, bold=True)
    W(ws, 4, 1, "CONCEPTO", bold=True)
    for i, mn in enumerate(MONTHS):
        W(ws, 4, 2+i, mn, bold=True, align="center")

    W(ws, P_LIQ_HDR, 1, "LIQUIDEZ", bold=True)
    for i in range(NB):
        r = P_LIQ_S + i
        W(ws, r, 1, BANKS[i])
        for mi, mn in enumerate(MONTHS):
            W(ws, r, 2+mi, f=f"=IFERROR({mn}!K{VB_S+i},0)", nf=EUR)
    W(ws, P_LIQ_T, 1, "TOTAL LIQUIDEZ", bold=True)
    for mi in range(12):
        mc = gcl(2+mi)
        W(ws, P_LIQ_T, 2+mi, f=f"=SUM({mc}{P_LIQ_S}:{mc}{P_LIQ_T-1})", nf=EUR, bold=True)

    W(ws, P_IA_HDR, 1, "INV. DINERO APORTADO", bold=True)
    for i in range(NI):
        r = P_IA_S + i
        W(ws, r, 1, INVS[i])
        for mi, mn in enumerate(MONTHS):
            W(ws, r, 2+mi, f=f"=IFERROR({mn}!K{VI_S+i},0)", nf=EUR)
    W(ws, P_IA_T, 1, "TOTAL APORTADO", bold=True)
    for mi in range(12):
        mc = gcl(2+mi)
        W(ws, P_IA_T, 2+mi, f=f"=SUM({mc}{P_IA_S}:{mc}{P_IA_T-1})", nf=EUR, bold=True)

    W(ws, P_IR_HDR, 1, "INV. VALOR REAL", bold=True)
    for i in range(NI):
        r = P_IR_S + i
        W(ws, r, 1, INVS[i])
        for mi, mn in enumerate(MONTHS):
            W(ws, r, 2+mi, f=f"=IFERROR({mn}!K{VR_S+i},0)", nf=EUR)
    W(ws, P_IR_T, 1, "TOTAL INV. REAL", bold=True)
    for mi in range(12):
        mc = gcl(2+mi)
        W(ws, P_IR_T, 2+mi, f=f"=SUM({mc}{P_IR_S}:{mc}{P_IR_T-1})", nf=EUR, bold=True)

    if ND > 0:
        W(ws, P_DAH, 1, "DEUDAS Y ACTIVOS", bold=True)
        for i in range(ND):
            r = P_DA_S + i
            W(ws, r, 1, f=f'=IFERROR(Configuracion!E{4+i},"")')
            W(ws, r, 2, f=f"=IFERROR(Configuracion!F{4+i},0)", nf=EUR)
            W(ws, r, 3, f=f"=IFERROR(Configuracion!G{4+i},0)", nf=EUR)
        W(ws, P_DA_T, 1, "ACTIVOS NETOS", bold=True)
        W(ws, P_DA_T, 2,
          f=f"=IFERROR(SUM(C{P_DA_S}:C{P_DA_T-1})-SUM(B{P_DA_S}:B{P_DA_T-1}),0)",
          nf=EUR, bold=True)

    W(ws, P_PAT_ROW, 1, "PATRIMONIO NETO", bold=True)
    for mi in range(12):
        mc = gcl(2+mi)
        if ND > 0:
            f = (f"=IFERROR({mc}{P_LIQ_T},0)+IFERROR({mc}{P_IR_T},0)"
                 f"+IFERROR(SUM(C{P_DA_S}:C{P_DA_T-1})-SUM(B{P_DA_S}:B{P_DA_T-1}),0)")
        else:
            f = f"=IFERROR({mc}{P_LIQ_T},0)+IFERROR({mc}{P_IR_T},0)"
        W(ws, P_PAT_ROW, 2+mi, f=f, nf=EUR, bold=True)

    CW(ws, 1, 30)
    for i in range(12): CW(ws, 2+i, 13)
    ws.freeze_panes = "B5"
    return ws


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# HOJA: Inversiones
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
def sh_inversiones(wb):
    ws = wb.create_sheet("Inversiones")
    W(ws, 1, 1, "INVERSIONES", bold=True)
    W(ws, 3, 1, "CAPITAL TOTAL INVERTIDO", bold=True)
    W(ws, 3, 2, f=f"=IFERROR(SUM(Configuracion!C4:C{3+NI}),0)", nf=EUR, bold=True)
    W(ws, 4, 1, "VALOR ACTUAL TOTAL", bold=True)
    W(ws, 4, 2,
      f=f"=IFERROR(LOOKUP(2,1/(PatrimonioTotal!B{P_IR_T}:M{P_IR_T}>0),PatrimonioTotal!B{P_IR_T}:M{P_IR_T}),0)",
      nf=EUR, bold=True)
    W(ws, 5, 1, "RENTABILIDAD EUR", bold=True)
    W(ws, 5, 2, f="=IFERROR(B4-B3,0)", nf=EUR, bold=True)
    W(ws, 6, 1, "RENTABILIDAD %", bold=True)
    W(ws, 6, 2, f="=IFERROR(B5/B3,0)", nf=PCT, bold=True)

    for c, lbl in enumerate(["TIPO","CAPITAL INICIAL","VALOR ACTUAL","RENT. EUR","RENT. %"], 1):
        W(ws, 9, c, lbl, bold=True)
    for i in range(NI):
        r = 10 + i
        pr = P_IR_S + i
        W(ws, r, 1, f=f'=IFERROR(Configuracion!B{4+i},"")')
        W(ws, r, 2, f=f"=IFERROR(Configuracion!C{4+i},0)", nf=EUR)
        W(ws, r, 3,
          f=f"=IFERROR(LOOKUP(2,1/(PatrimonioTotal!B{pr}:M{pr}>0),PatrimonioTotal!B{pr}:M{pr}),0)",
          nf=EUR)
        W(ws, r, 4, f=f"=IFERROR(C{r}-B{r},0)", nf=EUR)
        W(ws, r, 5, f=f"=IFERROR(D{r}/B{r},0)", nf=PCT)

    W(ws, 16, 1, "EVOLUCION MENSUAL", bold=True)
    for c, lbl in enumerate(["Mes","Valor Real Total","Capital Aportado","Rentabilidad"], 1):
        W(ws, 17, c, lbl, bold=True)
    for i, mn in enumerate(MONTHS):
        r = 18 + i
        mc = gcl(2+i)
        W(ws, r, 1, mn)
        W(ws, r, 2, f=f"=IFERROR(PatrimonioTotal!{mc}{P_IR_T},0)", nf=EUR)
        W(ws, r, 3, f=f"=IFERROR(PatrimonioTotal!{mc}{P_IA_T},0)", nf=EUR)
        W(ws, r, 4, f=f"=IFERROR(B{r}-C{r},0)", nf=EUR)

    for c, w in [(1,26),(2,16),(3,16),(4,14),(5,12)]: CW(ws, c, w)
    return ws


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# HOJA: Objetivos
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
def sh_objetivos(wb):
    ws = wb.create_sheet("Objetivos")
    W(ws, 1, 1, "OBJETIVOS PATRIMONIALES", bold=True)
    W(ws, 3, 1, "PATRIMONIO ACTUAL:", bold=True)
    W(ws, 3, 2,
      f=f"=IFERROR(LOOKUP(2,1/(PatrimonioTotal!B{P_PAT_ROW}:M{P_PAT_ROW}>0),PatrimonioTotal!B{P_PAT_ROW}:M{P_PAT_ROW}),0)",
      nf=EUR, bold=True)
    W(ws, 5, 1, "Edita los importes y anos cuando tengas claros tus objetivos")
    for c, lbl in enumerate(["OBJETIVO","ANO META","IMPORTE META","% COMPLETADO"], 1):
        W(ws, 6, c, lbl, bold=True)
    for i, (name, yr) in enumerate([("Objetivo 1 (editar)", 2027),
                                     ("Objetivo 2 (editar)", 2030),
                                     ("Objetivo 3 (editar)", 2035)]):
        r = 7 + i
        W(ws, r, 1, name); W(ws, r, 2, yr); W(ws, r, 3, 0, nf=EUR)
        W(ws, r, 4, f=f"=IFERROR(IF(C{r}>0,$B$3/C{r},0),0)", nf=PCT)

    CC = 14; CH = 20
    W(ws, CH, CC, "Mes", bold=True); W(ws, CH, CC+1, "Patrimonio Total", bold=True)
    for i, mn in enumerate(MONTHS):
        r = CH + 1 + i
        W(ws, r, CC, mn)
        W(ws, r, CC+1, f=f"=IFERROR(PatrimonioTotal!{gcl(2+i)}{P_PAT_ROW},0)", nf=EUR)

    chart = BarChart()
    chart.type = "col"; chart.title = "Evolucion Patrimonio Total"
    chart.y_axis.title = "EUR"; chart.x_axis.title = "Mes"
    chart.width = 18; chart.height = 11
    chart.add_data(Reference(ws, min_col=CC+1, max_col=CC+1, min_row=CH, max_row=CH+12),
                   titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=CC, max_col=CC, min_row=CH+1, max_row=CH+12))
    ws.add_chart(chart, "A12")
    for c, w in [(1,20),(2,10),(3,16),(4,14),(CC,12),(CC+1,16)]: CW(ws, c, w)
    return ws


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# HOJA: Presupuesto
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
def sh_presupuesto(wb):
    ws = wb.create_sheet("Presupuesto")
    W(ws, 1, 1, f"PRESUPUESTO ANUAL {YEAR}", bold=True)
    W(ws, 2, 1, "Columna REAL = suma automatica de los 12 meses")

    def real_f(cat_cell, rs, re):
        parts = [f"SUMIF({mn}!$C${rs}:$C${re},{cat_cell},{mn}!$E${rs}:$E${re})"
                 for mn in MONTHS]
        return "=IFERROR(" + "+".join(parts) + ",0)"

    IHR = 4
    for c, lbl in enumerate(["CATEGORIA","PRESUPUESTO","REAL (anual)","DIFERENCIA"], 1):
        W(ws, IHR, c, lbl, bold=True)
    for i in range(NIC):
        r = IHR + 1 + i
        W(ws, r, 1, f=f'=IFERROR(Configuracion!I{4+i},"")')
        W(ws, r, 2, 0, nf=EUR)
        W(ws, r, 3, f=real_f(f"A{r}", ING_S, ING_E), nf=EUR)
        W(ws, r, 4, f=f"=IFERROR(C{r}-B{r},0)", nf=EUR)
    ITR = IHR + 1 + NIC
    W(ws, ITR, 1, "TOTAL INGRESOS", bold=True)
    W(ws, ITR, 2, f=f"=SUM(B{IHR+1}:B{ITR-1})", nf=EUR, bold=True)
    W(ws, ITR, 3, f=f"=SUM(C{IHR+1}:C{ITR-1})", nf=EUR, bold=True)
    W(ws, ITR, 4, f=f"=IFERROR(C{ITR}-B{ITR},0)", nf=EUR, bold=True)

    GHR = ITR + 3
    for c, lbl in enumerate(["CATEGORIA","PRESUPUESTO","REAL (anual)","DIFERENCIA"], 1):
        W(ws, GHR, c, lbl, bold=True)
    for i in range(NEC):
        r = GHR + 1 + i
        W(ws, r, 1, f=f'=IFERROR(Configuracion!J{4+i},"")')
        W(ws, r, 2, 0, nf=EUR)
        W(ws, r, 3, f=real_f(f"A{r}", GAS_S, GAS_E), nf=EUR)
        W(ws, r, 4, f=f"=IFERROR(C{r}-B{r},0)", nf=EUR)
    GTR = GHR + 1 + NEC
    W(ws, GTR, 1, "TOTAL GASTOS", bold=True)
    W(ws, GTR, 2, f=f"=SUM(B{GHR+1}:B{GTR-1})", nf=EUR, bold=True)
    W(ws, GTR, 3, f=f"=SUM(C{GHR+1}:C{GTR-1})", nf=EUR, bold=True)
    W(ws, GTR, 4, f=f"=IFERROR(C{GTR}-B{GTR},0)", nf=EUR, bold=True)

    for c, w in [(1,24),(2,16),(3,16),(4,14)]: CW(ws, c, w)
    return ws


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# HOJA: ResumenMensual
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
def sh_resumen_mensual(wb):
    ws = wb.create_sheet("ResumenMensual")
    W(ws, 1, 1, "RESUMEN MENSUAL", bold=True)
    W(ws, 2, 1, "Mes seleccionado:", bold=True)
    W(ws, 2, 2, "Junio", bold=True)
    dv = DataValidation(type="list", formula1='"' + ",".join(MONTHS) + '"',
                        allow_blank=False, showErrorMessage=False)
    ws.add_data_validation(dv); dv.add(ws.cell(2, 2))

    W(ws, 4,  1, "TOTALES",         bold=True)
    W(ws, 5,  1, "Ingresos",        bold=True)
    W(ws, 5,  2, f=f'=IFERROR(INDIRECT(B2&"!E{ING_TOT}"),0)', nf=EUR, bold=True)
    W(ws, 6,  1, "Gastos",          bold=True)
    W(ws, 6,  2, f=f'=IFERROR(INDIRECT(B2&"!E{GAS_TOT}"),0)', nf=EUR, bold=True)
    W(ws, 7,  1, "Inversiones",     bold=True)
    W(ws, 7,  2, f=f'=IFERROR(INDIRECT(B2&"!E{INV_TOT}"),0)', nf=EUR, bold=True)
    W(ws, 8,  1, "Liquidez total",  bold=True)
    W(ws, 8,  2, f=f'=IFERROR(INDIRECT(B2&"!K{VT_LIQ}"),0)',  nf=EUR, bold=True)
    W(ws, 9,  1, "Balance",         bold=True)
    W(ws, 9,  2, f="=IFERROR(B5-B6,0)", nf=EUR, bold=True)
    W(ws, 10, 1, "Tasa de ahorro",  bold=True)
    W(ws, 10, 2, f=f'=IFERROR(INDIRECT(B2&"!H{H_SAV}"),0)', nf=PCT, bold=True)

    W(ws, 12, 1, "INGRESOS POR CATEGORIA", bold=True); W(ws, 12, 2, "Importe", bold=True)
    for i in range(NIC):
        r = 13 + i
        W(ws, r, 1, f=f'=IFERROR(Configuracion!I{4+i},"")')
        W(ws, r, 2, f=f'=IFERROR(INDIRECT(B2&"!H{DSG_INC_S+i}"),0)', nf=EUR)

    GS2 = 13 + NIC + 2
    W(ws, GS2, 1, "GASTOS POR CATEGORIA", bold=True); W(ws, GS2, 2, "Importe", bold=True)
    for i in range(NEC):
        r = GS2 + 1 + i
        W(ws, r, 1, f=f'=IFERROR(Configuracion!J{4+i},"")')
        W(ws, r, 2, f=f'=IFERROR(INDIRECT(B2&"!H{DSG_GAS_S+i}"),0)', nf=EUR)

    VA = GS2 + NEC + 3
    W(ws, VA,   1, "VALORES ACTUALES",         bold=True)
    W(ws, VA+1, 1, "Banco",                    bold=True)
    W(ws, VA+1, 2, "Saldo",                    bold=True)
    for i in range(NB):
        r = VA + 2 + i
        W(ws, r, 1, f=f'=IFERROR(Configuracion!A{4+i},"")')
        W(ws, r, 2, f=f'=IFERROR(INDIRECT(B2&"!K{VB_S+i}"),0)', nf=EUR)

    IV = VA + 2 + NB + 1
    W(ws, IV,   1, "Inversion (Valor Real)",   bold=True)
    W(ws, IV,   2, "Valor",                    bold=True)
    for i in range(NI):
        r = IV + 1 + i
        W(ws, r, 1, f=f'=IFERROR(Configuracion!B{4+i},"")')
        W(ws, r, 2, f=f'=IFERROR(INDIRECT(B2&"!K{VR_S+i}"),0)', nf=EUR)

    for c, w in [(1,26),(2,16)]: CW(ws, c, w)
    return ws


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# HOJA: ResumenAnual
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
def sh_resumen_anual(wb):
    ws = wb.create_sheet("ResumenAnual")
    W(ws, 1, 1, f"RESUMEN ANUAL {YEAR}", bold=True)

    def s12(row_ref):
        return "=IFERROR(" + "+".join(f"{mn}!E{row_ref}" for mn in MONTHS) + ",0)"

    W(ws, 3, 1, "TOTAL INGRESOS",     bold=True); W(ws, 3, 2, f=s12(ING_TOT), nf=EUR, bold=True)
    W(ws, 4, 1, "TOTAL GASTOS",       bold=True); W(ws, 4, 2, f=s12(GAS_TOT), nf=EUR, bold=True)
    W(ws, 5, 1, "TOTAL INVERSIONES",  bold=True); W(ws, 5, 2, f=s12(INV_TOT), nf=EUR, bold=True)
    W(ws, 6, 1, "BALANCE AÃ‘O",        bold=True); W(ws, 6, 2, f="=IFERROR(B3-B4,0)", nf=EUR, bold=True)
    W(ws, 7, 1, "TASA AHORRO MEDIA",  bold=True); W(ws, 7, 2, f="=IFERROR(B6/B3,0)", nf=PCT, bold=True)
    W(ws, 8, 1, "PATRIMONIO DIC.",    bold=True)
    W(ws, 8, 2, f=f"=IFERROR(PatrimonioTotal!M{P_PAT_ROW},0)", nf=EUR, bold=True)

    for c, lbl in enumerate(["MES","INGRESOS","GASTOS","INVERSIONES","BALANCE","TASA AHORRO","PATRIMONIO"], 1):
        W(ws, 11, c, lbl, bold=True)
    for i, mn in enumerate(MONTHS):
        r = 12 + i
        mc = gcl(2+i)
        W(ws, r, 1, mn)
        W(ws, r, 2, f=f"=IFERROR({mn}!E{ING_TOT},0)", nf=EUR)
        W(ws, r, 3, f=f"=IFERROR({mn}!E{GAS_TOT},0)", nf=EUR)
        W(ws, r, 4, f=f"=IFERROR({mn}!E{INV_TOT},0)", nf=EUR)
        W(ws, r, 5, f=f"=IFERROR(B{r}-C{r},0)",       nf=EUR)
        W(ws, r, 6, f=f"=IFERROR((B{r}-C{r})/B{r},0)",nf=PCT)
        W(ws, r, 7, f=f"=IFERROR(PatrimonioTotal!{mc}{P_PAT_ROW},0)", nf=EUR)

    W(ws, 26, 1, "INGRESOS POR CATEGORIA (anual)", bold=True); W(ws, 26, 2, "Importe", bold=True)
    for i in range(NIC):
        r = 27 + i
        W(ws, r, 1, f=f'=IFERROR(Configuracion!I{4+i},"")')
        parts = [f"SUMIF({mn}!$C${ING_S}:$C${ING_E},A{r},{mn}!$E${ING_S}:$E${ING_E})" for mn in MONTHS]
        W(ws, r, 2, f="=IFERROR("+"+".join(parts)+",0)", nf=EUR)

    GX = 27 + NIC + 2
    W(ws, GX, 1, "GASTOS POR CATEGORIA (anual)", bold=True); W(ws, GX, 2, "Importe", bold=True)
    for i in range(NEC):
        r = GX + 1 + i
        W(ws, r, 1, f=f'=IFERROR(Configuracion!J{4+i},"")')
        parts = [f"SUMIF({mn}!$C${GAS_S}:$C${GAS_E},A{r},{mn}!$E${GAS_S}:$E${GAS_E})" for mn in MONTHS]
        W(ws, r, 2, f="=IFERROR("+"+".join(parts)+",0)", nf=EUR)

    for c, w in [(1,28),(2,14),(3,14),(4,14),(5,14),(6,12),(7,14)]: CW(ws, c, w)
    return ws


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# HOJA: Dashboard
# Panel de control con todas las metricas clave
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
def sh_dashboard(wb):
    ws = wb.create_sheet("Dashboard")

    W(ws, 1, 1, f"DASHBOARD - FINANZAS {YEAR}", bold=True)
    W(ws, 2, 1, "Mes activo:", bold=True)
    W(ws, 2, 2, "Junio", bold=True)
    dv = DataValidation(type="list", formula1='"' + ",".join(MONTHS) + '"',
                        allow_blank=False, showErrorMessage=False)
    ws.add_data_validation(dv); dv.add(ws.cell(2, 2))

    # â”€â”€ BLOQUE 1: PATRIMONIO â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    W(ws, 4, 1, "PATRIMONIO", bold=True)
    W(ws, 5, 1, "Patrimonio neto total")
    W(ws, 5, 2,
      f=f"=IFERROR(LOOKUP(2,1/(PatrimonioTotal!B{P_PAT_ROW}:M{P_PAT_ROW}>0),PatrimonioTotal!B{P_PAT_ROW}:M{P_PAT_ROW}),0)",
      nf=EUR, bold=True)
    W(ws, 6, 1, "Liquidez (mes activo)")
    W(ws, 6, 2, f=f'=IFERROR(INDIRECT(B2&"!K{VT_LIQ}"),0)', nf=EUR)
    W(ws, 7, 1, "Inversiones - valor real")
    W(ws, 7, 2,
      f=f"=IFERROR(LOOKUP(2,1/(PatrimonioTotal!B{P_IR_T}:M{P_IR_T}>0),PatrimonioTotal!B{P_IR_T}:M{P_IR_T}),0)",
      nf=EUR)
    W(ws, 8, 1, "Inversiones - rentabilidad EUR")
    W(ws, 8, 2, f=f"=IFERROR(B7-SUM(Configuracion!C4:C{3+NI}),0)", nf=EUR)
    W(ws, 9, 1, "Inversiones - rentabilidad %")
    W(ws, 9, 2, f=f"=IFERROR(B8/SUM(Configuracion!C4:C{3+NI}),0)", nf=PCT)

    # â”€â”€ BLOQUE 2: ESTE MES â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    W(ws, 11, 1, "ESTE MES", bold=True)
    W(ws, 12, 1, "Ingresos")
    W(ws, 12, 2, f=f'=IFERROR(INDIRECT(B2&"!E{ING_TOT}"),0)', nf=EUR)
    W(ws, 13, 1, "Gastos")
    W(ws, 13, 2, f=f'=IFERROR(INDIRECT(B2&"!E{GAS_TOT}"),0)', nf=EUR)
    W(ws, 14, 1, "Inversiones aportadas")
    W(ws, 14, 2, f=f'=IFERROR(INDIRECT(B2&"!E{INV_TOT}"),0)', nf=EUR)
    W(ws, 15, 1, "Balance (Ingresos - Gastos)", bold=True)
    W(ws, 15, 2, f="=IFERROR(B12-B13,0)", nf=EUR, bold=True)
    W(ws, 16, 1, "Tasa de ahorro", bold=True)
    W(ws, 16, 2, f=f'=IFERROR(INDIRECT(B2&"!H{H_SAV}"),0)', nf=PCT, bold=True)

    # â”€â”€ BLOQUE 3: LIQUIDEZ POR CUENTA â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    r = 18
    W(ws, r, 1, "LIQUIDEZ POR CUENTA", bold=True)
    for i in range(NB):
        W(ws, r+1+i, 1, f=f'=IFERROR(Configuracion!A{4+i},"")')
        W(ws, r+1+i, 2, f=f'=IFERROR(INDIRECT(B2&"!K{VB_S+i}"),0)', nf=EUR)
    W(ws, r+1+NB, 1, "TOTAL", bold=True)
    W(ws, r+1+NB, 2, f=f"=IFERROR(SUM(B{r+1}:B{r+NB}),0)", nf=EUR, bold=True)
    RLIQ_TOT = r + 1 + NB  # fila total liquidez en dashboard

    # â”€â”€ BLOQUE 4: INVERSIONES POR TIPO â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    r = 18 + NB + 4
    W(ws, r,   1, "INVERSIONES",   bold=True)
    W(ws, r,   2, "Aportado",      bold=True)
    W(ws, r,   3, "Valor real",    bold=True)
    W(ws, r,   4, "Rent. %",       bold=True)
    for i in range(NI):
        rr = r + 1 + i
        W(ws, rr, 1, f=f'=IFERROR(Configuracion!B{4+i},"")')
        W(ws, rr, 2, f=f'=IFERROR(INDIRECT(B2&"!K{VI_S+i}"),0)', nf=EUR)
        W(ws, rr, 3, f=f'=IFERROR(INDIRECT(B2&"!K{VR_S+i}"),0)', nf=EUR)
        W(ws, rr, 4, f=f"=IFERROR((C{rr}-B{rr})/B{rr},0)", nf=PCT)

    # â”€â”€ BLOQUE 5: FONDO DE EMERGENCIA â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    r = 18 + NB + 4 + NI + 3
    W(ws, r,   1, "FONDO DE EMERGENCIA", bold=True)
    W(ws, r+1, 1, "Gastos este mes")
    W(ws, r+1, 2, f=f'=IFERROR(INDIRECT(B2&"!E{GAS_TOT}"),0)', nf=EUR)
    W(ws, r+2, 1, "Tu liquidez total")
    W(ws, r+2, 2, f=f"=IFERROR(B{RLIQ_TOT},0)", nf=EUR)
    W(ws, r+3, 1, "Meses cubiertos", bold=True)
    W(ws, r+3, 2, f=f"=IFERROR(B{r+2}/B{r+1},0)", nf="0.0", bold=True)
    W(ws, r+3, 3, "(recomendado: minimo 3 meses)")

    # â”€â”€ BLOQUE 6: OBJETIVOS â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    r2 = r + 6
    W(ws, r2,   1, "OBJETIVOS", bold=True)
    W(ws, r2,   2, "Meta",      bold=True)
    W(ws, r2,   3, "% logrado", bold=True)
    for i in range(3):
        rr = r2 + 1 + i
        W(ws, rr, 1, f=f'=IFERROR(Objetivos!A{7+i},"")')
        W(ws, rr, 2, f=f"=IFERROR(Objetivos!C{7+i},0)", nf=EUR)
        W(ws, rr, 3, f=f"=IFERROR(Objetivos!D{7+i},0)", nf=PCT)

    # â”€â”€ GRAFICO EVOLUCION (col F-G como datos, grafico a la derecha) â”€â”€
    CC = 6; CH = 4
    W(ws, CH,   CC, "Mes",       bold=True)
    W(ws, CH,   CC+1, "Patrimonio", bold=True)
    for i, mn in enumerate(MONTHS):
        rr = CH + 1 + i
        W(ws, rr, CC,   mn)
        W(ws, rr, CC+1, f=f"=IFERROR(PatrimonioTotal!{gcl(2+i)}{P_PAT_ROW},0)", nf=EUR)

    chart = BarChart()
    chart.type = "col"; chart.title = "Evolucion del Patrimonio 2025"
    chart.y_axis.title = "EUR"; chart.width = 18; chart.height = 14
    chart.add_data(Reference(ws, min_col=CC+1, max_col=CC+1, min_row=CH, max_row=CH+12),
                   titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=CC, max_col=CC, min_row=CH+1, max_row=CH+12))
    ws.add_chart(chart, "H4")

    for c, w in [(1,30),(2,16),(3,16),(4,20),(5,2),(CC,10),(CC+1,14)]: CW(ws, c, w)
    ws.freeze_panes = "A3"

    # Mover Dashboard a primera posicion
    wb.move_sheet("Dashboard", offset=-wb.sheetnames.index("Dashboard"))
    return ws


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# MAIN
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
def build():
    wb = openpyxl.Workbook()
    sh_configuracion(wb)
    sh_patrimonio(wb)
    sh_inversiones(wb)
    sh_objetivos(wb)
    sh_presupuesto(wb)
    sh_resumen_mensual(wb)
    for mn in MONTHS: sh_mes(wb, mn)
    sh_resumen_anual(wb)
    sh_dashboard(wb)   # crea + mueve a posicion 0

    # Pre-rellenar Junio con saldos reales a 29/06/2025
    ws_jun = wb["Junio"]
    for i, v in enumerate(JUNIO_BANCOS):
        ws_jun.cell(VB_S+i, 11).value = v
        ws_jun.cell(VB_S+i, 11).number_format = EUR
    for i, v in enumerate(JUNIO_INV_AP):
        ws_jun.cell(VI_S+i, 11).value = v
        ws_jun.cell(VI_S+i, 11).number_format = EUR
    for i, v in enumerate(JUNIO_INV_RE):
        ws_jun.cell(VR_S+i, 11).value = v
        ws_jun.cell(VR_S+i, 11).number_format = EUR

    out = "Plantilla_Finanzas_2025_v5.xlsx"
    wb.save(out)
    print(f"OK: {out}")
    print("Orden hojas: " + " | ".join(ws.title for ws in wb.worksheets))

if __name__ == "__main__":
    build()
